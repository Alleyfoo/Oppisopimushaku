from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from apprscan.jobs.storage import write_master_workbook


def read_headers(ws):
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def test_master_workbook_creation(tmp_path: Path):
    shortlist = pd.DataFrame({"business_id": ["123", "456"], "score": [10, 8]})
    excluded = pd.DataFrame({"business_id": ["999"], "excluded_reason": ["test"]})
    jobs_all = pd.DataFrame({"job_url": ["https://a/1", "https://a/2"], "company_domain": ["a.com", "a.com"]})
    jobs_new = pd.DataFrame({"job_url": ["https://a/2"]})
    crawl_stats = pd.DataFrame({"domain": ["a.com"], "skip_reason": [None]})
    activity = pd.DataFrame({"business_id": ["123"], "job_count_total": [2], "recruiting_active": [True]})

    out_path = tmp_path / "master.xlsx"
    write_master_workbook(
        out_path,
        shortlist=shortlist,
        excluded=excluded,
        jobs_all=jobs_all,
        jobs_new=jobs_new,
        crawl_stats=crawl_stats,
        activity=activity,
    )

    assert out_path.exists()
    wb = load_workbook(out_path, read_only=True)
    sheets = wb.sheetnames
    assert "Shortlist" in sheets
    assert "Jobs_All" in sheets
    assert "Jobs_New" in sheets
    assert "Crawl_Stats" in sheets
    assert "Company_Activity" in sheets
    ws = wb["Shortlist"]
    assert "business_id" in read_headers(ws)
    ws_jobs = wb["Jobs_All"]
    assert "job_url" in read_headers(ws_jobs)
